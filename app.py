import os
import uuid
import datetime
import pandas as pd
from flask import Flask, request, jsonify, send_file, make_response
from parser import parse_system_info

app = Flask(__name__, static_folder='public', static_url_path='')

DATA_FILE = 'assets.json'

def load_assets():
    if not os.path.exists(DATA_FILE):
        return []
    try:
        import json
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []

def save_assets(assets):
    try:
        import json
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(assets, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving assets: {e}")

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/assets', methods=['GET'])
def get_assets():
    return jsonify(load_assets())

@app.route('/api/upload', methods=['POST'])
def upload_log():
    log_content = ""
    
    # Check if file was uploaded via multipart/form-data
    if 'file' in request.files:
        file = request.files['file']
        if file.filename != '':
            log_content = file.read().decode('utf-8', errors='ignore')
            
    # Check if raw data was posted (e.g. curl -d @file)
    if not log_content and request.data:
        log_content = request.data.decode('utf-8', errors='ignore')
        
    # Check if it was sent as form text parameter
    if not log_content and request.form.get('log_text'):
        log_content = request.form.get('log_text')
        
    # Check if it was sent as JSON
    if not log_content and request.is_json:
        log_content = request.json.get('log_text', '')

    if not log_content or not log_content.strip():
        return jsonify({"success": False, "error": "No log content provided"}), 400

    try:
        parsed_data = parse_system_info(log_content)
        
        # Verify if we could parse at least a hostname
        if not parsed_data["hostname"]:
            return jsonify({"success": False, "error": "Could not parse hostname from log"}), 400

        assets = load_assets()
        
        # Check if asset with this hostname already exists
        existing_idx = -1
        for idx, asset in enumerate(assets):
            if asset.get("hostname").lower() == parsed_data["hostname"].lower():
                existing_idx = idx
                break
                
        # Fill in metadata
        parsed_data["last_updated"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        parsed_data["sender_ip"] = request.remote_addr
        if not parsed_data.get("ip_address"):
            parsed_data["ip_address"] = request.remote_addr
        
        if existing_idx != -1:
            # Preserve the existing ID
            parsed_data["id"] = assets[existing_idx]["id"]
            assets[existing_idx] = parsed_data
            status = "updated"
        else:
            parsed_data["id"] = uuid.uuid4().hex
            assets.append(parsed_data)
            status = "created"
            
        save_assets(assets)
        return jsonify({"success": True, "status": status, "id": parsed_data["id"], "hostname": parsed_data["hostname"]})
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Error parsing log: {str(e)}"}), 500

@app.route('/api/assets/<asset_id>', methods=['DELETE'])
def delete_asset(asset_id):
    assets = load_assets()
    new_assets = [a for a in assets if a.get("id") != asset_id]
    if len(assets) == len(new_assets):
        return jsonify({"success": False, "error": "Asset not found"}), 404
        
    save_assets(new_assets)
    return jsonify({"success": True, "message": "Asset deleted successfully"})

@app.route('/api/assets/clear', methods=['POST'])
def clear_assets():
    save_assets([])
    return jsonify({"success": True, "message": "All assets cleared successfully"})

@app.route('/api/download-bat', methods=['GET'])
def download_bat():
    if os.path.exists('collect_info.bat'):
        return send_file('collect_info.bat', as_attachment=True, download_name='collect_info.bat')
    return make_response("Batch file not found", 404)

@app.route('/api/export', methods=['GET'])
def export_excel():
    assets = load_assets()
    if not assets:
        return make_response("No data available to export", 400)
        
    # Transform list of dictionaries into a flat list for DataFrame
    flat_data = []
    for asset in assets:
        # Create user profiles summary
        profiles = [f"{p['local_path']} ({p['sid']})" for p in asset.get("configured_user_profiles", [])]
        profiles_str = "\n".join(profiles)
        
        # Create user directories summary
        user_dirs_str = ", ".join(asset.get("user_directories", []))
        
        # Create network adapters summary
        adapters = [f"{a['name']}: {a['status']} ({a['mac_address']})" for a in asset.get("network_adapters", [])]
        adapters_str = "\n".join(adapters)
        
        flat_data.append({
            "Hostname": asset.get("hostname", ""),
            "Serial Number": asset.get("serial_number", ""),
            "Model Name": asset.get("model_name", ""),
            "IP Address": asset.get("ip_address", ""),
            "User Directories": user_dirs_str,
            "User Profiles Count": len(asset.get("configured_user_profiles", [])),
            "Configured User Profiles Detail": profiles_str,
            "Trend Micro Agent Version": asset.get("trend_micro_agent_version", ""),
            "Trend Micro Scan Engine": asset.get("trend_micro_scan_engine", ""),
            "Virus Scan Engine (wide-char)": asset.get("virus_scan_engine_alt", ""),
            "SelfScan Installed": "Yes" if asset.get("self_scan_installed", False) else "No",
            "SelfScan Path": asset.get("self_scan_path", ""),
            "SelfScan Command": asset.get("self_scan_command", ""),
            "Netskope Installed": "Yes" if asset.get("netskope_installed", False) else "No",
            "Network Adapters Count": len(asset.get("network_adapters", [])),
            "Network Adapters Detail": adapters_str,
            "Last Updated": asset.get("last_updated", ""),
            "Sender IP": asset.get("sender_ip", "")
        })
        
    df = pd.DataFrame(flat_data)
    
    # Export to Excel
    temp_file = "assets_temp.xlsx"
    
    try:
        # We can format the Excel output using openpyxl via pandas
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='System Assets', index=False)
            
            # Auto-adjust columns width and style header
            workbook = writer.book
            worksheet = writer.sheets['System Assets']
            
            # Format header row to look premium
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            cell_font = Font(name="Segoe UI", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            
            # Style header
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                
            # Style data cells and auto-adjust widths
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = thin_border
                    # Align number columns centered, long texts left
                    if col in [1, 2, 3, 4, 6, 8, 9, 10, 11, 14, 15, 17, 18]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
            
            # Set row heights
            worksheet.row_dimensions[1].height = 28
            for r in range(2, len(df) + 2):
                worksheet.row_dimensions[r].height = 20
                
            # Set col widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        # count longest line if wraptext is used
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            if len(line) > max_len:
                                max_len = len(line)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
                
        # Send file to client
        return_response = send_file(
            temp_file, 
            as_attachment=True, 
            download_name=f"Asset_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Delete file after sending (via background task if needed, but since it's simple we can just return and delete)
        # Actually, in Flask we can delete it in a teardown or just let it overwrite next time. Overwriting is perfectly fine.
        return return_response
        
    except Exception as e:
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass
        return make_response(f"Excel Generation Failed: {str(e)}", 500)

if __name__ == '__main__':
    # Create public folder if not exists
    os.makedirs('public', exist_ok=True)
    # Start app
    app.run(host='0.0.0.0', port=3000, debug=True)
